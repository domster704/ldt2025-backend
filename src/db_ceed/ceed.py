from itertools import chain
from os import PathLike

from mimesis import Person, Gender
from mimesis.locales import Locale
from openpyxl.reader.excel import load_workbook
from sqlalchemy import create_engine, text


def open_excel_doc(file_path: PathLike):
    wb = load_workbook(file_path)
    ws = wb.active
    iter_rows = ws.iter_rows(values_only=False)
    row = next(iter_rows)
    i = 0

    def validate_float_cell(value: str):
        if value != '–' and value != '-' and value is not None:
            if not isinstance(value, str):
                return float(value)
            else:
                return float(value.replace(',', '.'))
        else:
            return None

    for row in ws.iter_rows(values_only=True):
        if i < 2:
            i += 1
            continue
        if row[0] is None:
            continue
        if isinstance(row[0], int):
            yield {
                "patient_id": row[0] if 'regular' not in file_path else row[0]+50,
                "diagnosis": row[1],
                'blood_gas_ph': validate_float_cell(row[2]),
                'blood_gas_co2': validate_float_cell(row[3]),
                'blood_gas_glu': validate_float_cell(row[4]),
                'blood_gas_lac': validate_float_cell(row[5]),
                'blood_gas_be': validate_float_cell(row[6]),
            }
        else:
            ids = row[0].split(', ')
            for id in ids:
                yield {
                    "patient_id": int(id) if 'regular' not in file_path else int(id)+50,
                    "diagnosis": row[1],
                    'blood_gas_ph': validate_float_cell(row[2]),
                    'blood_gas_co2': validate_float_cell(row[3]),
                    'blood_gas_glu': validate_float_cell(row[4]),
                    'blood_gas_lac': validate_float_cell(row[5]),
                    'blood_gas_be': validate_float_cell(row[6]),
                }

def db_seed():
    """ Заносим данные в БД """
    engine = create_engine("sqlite:///./../../app.db")

    p = Person(Locale.RU)
    # Заносим пациентов
    for i in range(210):
        full_name = p.full_name(gender=Gender.FEMALE)
        with engine.begin() as conn:
            conn.execute(
                text("INSERT INTO patients (full_name) VALUES (:full_name)"),
                {"full_name": full_name}
            )

    # Заносим базовую информацию
    for data in chain(open_excel_doc("./static/hypoxia.xlsx"), open_excel_doc("./static/regular.xlsx")):
        with engine.begin() as conn:
            conn.execute(
                text("INSERT INTO patient_info (patient_id, diagnosis, blood_gas_ph, blood_gas_co2, blood_gas_glu, blood_gas_lac, blood_gas_be)"
                     "VALUES (:patient_id, :diagnosis, :blood_gas_ph, :blood_gas_co2, :blood_gas_glu, :blood_gas_lac, :blood_gas_be)"),
                data
            )
db_seed()